# utils.py
import os
import pickle
import re
import json
import csv
from PyPDF2 import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from tqdm import tqdm

def list_docs(folder="data"):
    exts = (".pdf", ".txt", ".md", ".csv", ".json", ".docx", ".xlsx", ".html", ".doc")
    files = []
    for root, _, filenames in os.walk(folder):
        for fn in filenames:
            if fn.lower().endswith(exts):
                files.append(os.path.join(root, fn))
    return files

def load_pdf(path):
    pages = []
    try:
        r = PdfReader(path)
        for p in r.pages:
            text = p.extract_text() or ""
            pages.append(text)
        return "\n".join(pages)
    except Exception as e:
        print(f"Error reading PDF {path}: {e}")
        return ""

def load_text_file(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception as e:
        print(f"Error reading text file {path}: {e}")
        return ""

def load_csv(path):
    try:
        rows = []
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(" ".join(row))
        return "\n".join(rows)
    except Exception as e:
        print(f"Error reading CSV {path}: {e}")
        return ""

def load_json(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            data = json.load(f)
        return json.dumps(data, indent=2)
    except Exception as e:
        print(f"Error reading JSON {path}: {e}")
        return ""

def load_docx(path):
    try:
        doc = DocxDocument(path)
        paragraphs = [p.text for p in doc.paragraphs]
        return "\n".join(paragraphs)
    except Exception as e:
        print(f"Error reading DOCX {path}: {e}")
        return ""

def load_xlsx(path):
    try:
        wb = load_workbook(path)
        content = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            content.append(f"Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                content.append(" ".join(str(v) if v is not None else "" for v in row))
        return "\n".join(content)
    except Exception as e:
        print(f"Error reading XLSX {path}: {e}")
        return ""

def load_html(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            html_content = f.read()
        soup = BeautifulSoup(html_content, "html.parser")
        # Remove script and style elements
        for script in soup(["script", "style"]):
            script.decompose()
        text = soup.get_text()
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        text = "\n".join(chunk for chunk in chunks if chunk)
        return text
    except Exception as e:
        print(f"Error reading HTML {path}: {e}")
        return ""

def load_document(path):
    """Load document based on file extension."""
    path_lower = path.lower()
    
    if path_lower.endswith(".pdf"):
        return load_pdf(path)
    elif path_lower.endswith(".csv"):
        return load_csv(path)
    elif path_lower.endswith(".json"):
        return load_json(path)
    elif path_lower.endswith(".docx"):
        return load_docx(path)
    elif path_lower.endswith(".xlsx"):
        return load_xlsx(path)
    elif path_lower.endswith(".html"):
        return load_html(path)
    elif path_lower.endswith((".txt", ".md")):
        return load_text_file(path)
    else:
        return load_text_file(path)

def chunk_text(text, chunk_size=800, overlap=150):
    """Split text into overlapping chunks with proper stride."""
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= chunk_size:
        return [text] if text else []

    chunks = []
    start = 0
    stride = chunk_size - overlap

    while start < len(text):
        end = min(start + chunk_size, len(text))
        chunk = text[start:end]
        if chunk.strip():
            chunks.append(chunk)
        
        # Ensure forward progress
        if end >= len(text):
            break
        start += stride

    return chunks if chunks else [text]

def save_pickle(obj, path):
    with open(path, "wb") as f:
        pickle.dump(obj, f)

def load_pickle(path):
    with open(path, "rb") as f:
        return pickle.load(f)
